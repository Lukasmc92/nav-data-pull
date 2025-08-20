# nav_data_pull_streamlit.py
import streamlit as st
from datetime import datetime, timedelta
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
import requests
from io import BytesIO
from tqdm import tqdm

# --- Page config ---
st.set_page_config(page_title="NAV Data Pull", layout="wide")

st.title("📊 Closed-End Fund Data Research")

# --- Download Tickers file ---
TICKERS_URL = "https://github.com/Lukasmc92/NAV-Tickers/raw/refs/heads/main/Tickers.xlsx"

@st.cache_data
def load_tickers():
    r = requests.get(TICKERS_URL)
    return pd.read_excel(BytesIO(r.content), engine="openpyxl")

df_tickers = load_tickers()
df_tickers = df_tickers.dropna(subset=["Fund", "NAV"])

fund_tickers = df_tickers["Fund"].tolist()
nav_tickers = df_tickers["NAV"].tolist()
fund_types = df_tickers["Fund Type"].tolist()
fund_subcats = df_tickers["Subcategory"].tolist()
fund_broadcats = df_tickers["Broad Category"].tolist()
fund_regions = df_tickers["Geographic Focus"].tolist()

# --- Date Picker ---
target_date = st.date_input(
    "Valuation Date (or last weekday before valuation date)",
    value=datetime.today(),
)

# Helper function to get close price
def get_close_price(ticker, date, start, end):
    try:
        data = yf.Ticker(ticker).history(start=start, end=end, auto_adjust=False)
        if data.empty or date not in data.index.strftime('%Y-%m-%d'):
            return None
        data.index = data.index.strftime('%Y-%m-%d')
        return data.loc[date, "Close"]
    except Exception as e:
        print(f"Skipping {ticker} due to error: {e}")
        return None

# Helper function to get fundamentals as of a specific date
def get_fundamentals_asof(ticker: str, as_of_date: str, quarterly=True):
    t = yf.Ticker(ticker)
    balance = t.quarterly_balance_sheet if quarterly else t.balance_sheet

    if balance.empty:
        return {"shares_outstanding": None, "total_debt": None, "outside equity": None, "report_date": None}

    as_of_date = pd.Timestamp(as_of_date)
    valid_dates = [d for d in balance.columns if d <= as_of_date]

    if not valid_dates:
        return {"shares_outstanding": None, "total_debt": None, "outside equity": None, "report_date": None}

    latest = max(valid_dates)

    shares = None
    debt = None
    otequity = 0

    for row in ["Ordinary Shares Number", "Share Issued"]:
        if row in balance.index:
            shares = balance.loc[row, latest]
            break

    for row in ["Total Debt", "Long Term Debt", "Current Debt"]:
        if row in balance.index:
            debt = balance.loc[row, latest]
            break

    for row in ["Preferred Securities Outside Stock Equity"]:
        if row in balance.index:
            otequity = balance.loc[row, latest]
            break

    return {
        "shares_outstanding": shares,
        "total_debt": debt,
        "outside equity": otequity,
        "report_date": latest
    }


# --- Run Button ---
if st.button("Download NAV Data"):
    start_date = (target_date - timedelta(days=2)).strftime('%Y-%m-%d')
    end_date = (target_date + timedelta(days=2)).strftime('%Y-%m-%d')
    date_str = target_date.strftime('%Y-%m-%d')

    rows = []
    progress_bar = st.progress(0)

    for idx, (fund, nav, types, subcategories, broadcats, regions) in enumerate(zip(fund_tickers, nav_tickers, fund_types, fund_subcats, fund_broadcats, fund_regions)):
        ticker_obj = yf.Ticker(fund)
        info = ticker_obj.info

        fund_name = info.get("longName", fund)
        
        # Get fundamentals as of the target date
        fundamentals = get_fundamentals_asof(fund, target_date)
        
        shares_outstanding = fundamentals.get("shares_outstanding")
        total_debt = fundamentals.get("total_debt")
        outside_equity = fundamentals.get("outside equity")  # Optional if you want to use it
        
        shares_millions = round(shares_outstanding / 1_000_000, 2) if shares_outstanding else None
        debt_millions = round(total_debt / 1_000_000, 2) if total_debt else None


        fund_price = get_close_price(fund, date_str, start_date, end_date)
        nav_price = get_close_price(nav, date_str, start_date, end_date)
        discount = fund_price / nav_price if fund_price and nav_price else None

        rows.append([
            fund_name, broadcats, types, subcategories, regions, date_str, fund, fund_price,
            nav, nav_price, discount, shares_millions, debt_millions
        ])

        progress_bar.progress((idx+1) / len(fund_tickers))

    df = pd.DataFrame(rows, columns=[
        "Fund Name", "Broad Category", "Fund Type", "Subcategory", "Geographic Focus", "Date", "Fund Ticker", "Fund Close Price",
        "NAV Ticker", "NAV Close Price", "Discount",
        "Shares Outstanding(M)", "Total Debt(M)"
    ])

    # Save to Excel
    excel_filename = f'Closed_End_Fund_Data_{date_str}.xlsx'
    df.to_excel(excel_filename, index=False, sheet_name='Sheet1')

    wb = load_workbook(excel_filename)
    ws = wb['Sheet1']
    message_row = ws.max_row + 2
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    method = "This file was created using python, streamlit, and yfinance to pull NAV pricing."
    ws.cell(row=message_row, column=1, value=f"Downloaded on {timestamp}. Method: {method}")
    wb.save(excel_filename)

    st.success("✅ NAV Data Pull Complete")
    st.dataframe(df)

    # Download button
    with open(excel_filename, "rb") as f:
        st.download_button(
            label="📥 Download Excel File",
            data=f,
            file_name=excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )










